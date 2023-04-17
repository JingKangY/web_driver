# -*- coding: utf-8 -*-
'''
@File    :   get_data.py
@Time    :   2022/7/19 8:30
@Author  :   YangJingKang(靖康阳)
@Gitee   :   https://gitee.com/jingkangyang/
@Jianshu :   靖康阳
@Contact :   1097904758@qq.com
@License :   (C)Copyright 2022-2025, Micro-Circle
@Desc    :   文件内容描述
'''
import openpyxl

from data.get_filepath import GetFilePath
from loging.log import *


class GetDataExcel:

    def read_all_excel(self, excelpath, sheetname):
        """
        获取所有用例
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :return: 返回表格所有数据
        """
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        all_list = []
        for row in range(2, self.sheet.max_row + 1):
            row_list = []
            for col in range(1, self.sheet.max_column + 1):
                row_list.append(self.sheet.cell(row, col).value)
            all_list.append(row_list)
        # print(all_list)

        log.info(f'成功获取所有用例，共{(self.sheet.max_row - 1)}条用例')
        log.info(all_list)
        return all_list

    def read_screen_excel(self, excelpath, sheetname, screen_rows: list):
        '''
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param screen_rows: 指定不需要读取的行数，传入为列表
        :return:返回筛选完成表格所有数据
        '''
        self.wb = openpyxl.load_workbook(excelpath)
        self.sheet = self.wb[sheetname]
        screen_list = []
        for row in range(2, self.sheet.max_row + 1):
            if row in screen_rows:
                pass
            else:
                row_list = []
                for col in range(1, self.sheet.max_column + 1):
                    row_list.append(self.sheet.cell(row, col).value)
                screen_list.append(row_list)
        # print(all_list)
        log.info(f'成功获取所有用例，共{(self.sheet.max_row - 1 - len(screen_rows))}条用例')
        log.info(screen_list)
        return screen_list

    def read_one_row_excel(self, excelpath, sheetname, number: int):
        """
        获取某一条测试用例
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param number: 索引 --> 从第二行开始为0的索引
        :return: 返回一行数据
        """
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        fist_list = []
        second_list = []
        for i in self.sheet[number + 1]:
            fist_list.append(i.value)
        second_list.append(fist_list)
        log.info(f'成功获取一条测试用例：{second_list}')
        return second_list

    def read_cell_excel(self, excelpath, sheetname, row: int, col: int):
        """
        读取具体某一行某一列数据（单元格）
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param row: 行号 --> int
        :param col: 列 --> int
        :return: 单元格数据
        """
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        result = self.sheet.cell(row + 1, col + 1).value  # 输出具体单元格数据
        log.info(f'数据读取成功： {result}')
        return result

    def read_some_rows_excel(self, excelpath, sheetname, row1: int, row2: int):
        """
        读取区间用例，所有列（例：取5-10行）
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param row1: 行号 --> int
        :param row2: 行号 --> int
        :return: 返回区间数据
        """
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        all_list = []
        for row in range(row1 + 1, row2 + 2):
            row_list = []
            for cols in range(1, self.sheet.max_column + 1):
                row_list.append(self.sheet.cell(row, cols).value)
            all_list.append(row_list)
        log.info(f'成功从获取第{row1}到{row2}用例,共{row2 - row1 + 1}条用例')
        log.info(all_list)
        return all_list

    def read_special_list_excel(self, excelpath, sheetname, datas: list, testname: str):
        '''
        指定列数据读取（例如：1、3、5、6）
        :param excelpath:  用例位置
        :param sheetname: 文件sheet名称
        :param datas: 需要使用的列，的列传入数字即可
        :param testname: 用例所属模块
        :return: 用例数据
        '''
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        cases_num = self.sheet.max_row
        list_cases = []
        for row in range(2, cases_num + 1):
            list_case = []
            for data in datas:
                dict_case = self.sheet.cell(row=row, column=data).value
                list_case.append(dict_case)
            list_cases.append(list_case)
        log.info(f'成功获取{testname}所有用例，共{cases_num - 1}条用例')
        log.info(list_cases)
        return list_cases

    def read_special_rows_list_excel(self, excelpath, sheetname, row1, row2, datas: list):
        '''
        指定列指定行读取数据
        :param excelpath: 用例位置
        :param sheetname: 文件sheet名称
        :param row1: 开始第几行（包含）
        :param row2: 结束第几行（包含）
        :param datas: 指定列读取数据（传入列表即可）
        :return: 返回用例
        '''
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        all_list = []
        for row in range(row1 + 1, row2 + 2):
            list_case = []
            for data in datas:
                dict_case = self.sheet.cell(row=row, column=data).value
                list_case.append(dict_case)
            all_list.append(list_case)
        log.info(f'成功从获取第{row1}到{row2}用例,共{row2 - row1 + 1}条用例')
        log.info(all_list)
        return all_list

    def write_excel(self, excelpath, sheetname, index, writevalue):
        """
        写入数据
        :param excelpath: 文件路径 -->  Excel
        :param sheetname: 文件sheet名称
        :param writevalue: 写入的数据
        :param index: 单元格位置
        :return: inster
        """
        self.wb = openpyxl.load_workbook(excelpath)  # 加载工作簿
        self.sheet = self.wb[sheetname]  # 获得sheet对象
        self.active = self.wb.active  # 写入文件时需要激活
        self.active[index] = writevalue  # 通过单元格写入 例如：A1 A2 A3等
        self.wb.save(excelpath)  # 保存
        log.info(f'成功在%s表的%s页%s单元格写入数据内容为：%s' % (excelpath, sheetname, index, writevalue))


read = GetDataExcel()

if __name__ == '__main__':
    # 实例化
    # read.read_all_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'), 'wa登录')  # 第一个参数是文档位置，第二个是sheet页的名称
    # read.read_screen_excel(GetFilePath().get_all_file_path(r'data/api_login1.xlsx'), 'login', [2, 3, 4, 10, 13])  # 指定部分行数不读取
    # read.read_special_list_excel(GetFilePath().get_all_file_path(r'data/api_login1.xlsx'), 'login', [1, 2, 3, 4, 5, 8, 13], '登录login')
    # read.read_special_list_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'), '登录模块', [1, 2, 3, 4, 5, 10, 15, 16, 17, 18], '登录接口')
    # read.read_one_row_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'), '登录模块', 1)
    # # read.read_some_rows_excel(GetFilePath().get_all_file_path(r'data/api_login1.xlsx'), 'login', 0, 7)
    # read.read_cell_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'), '登录模块', 1, 1)
    # read.read_special_rows_list_excel(GetFilePath().get_all_file_path(r'data/双创项目接口文档.xlsx'), '登录模块', 1, 1, [1, 2, 3])

    # read.read_special_rows_list_excel(GetFilePath().get_all_file_path(r'data/api_login1.xlsx'),
    #   read_cell_excel                                'login', 5, 7, [10, 14])
    # case_info = read.read_special_rows_list_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'),
    #                                               '二行情诗-点赞列表', 1, 1, [2, 4, 7, 8, 9, 11, 12, 13, 14, 15, 16])
    #
    # title, desc, url, method, request_type, args, request_head, step, expect_code, expect_mes, expect_total_count = \
    # case_info[0]
    # print(title)
    # read.read_special_list_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'),
    #                              '二行情诗-发送弹幕', [2, 4, 7, 8, 9, 11, 12, 13, 14, 15, 16],
    #                              '发送弹幕')
    read.read_special_rows_list_excel(GetFilePath().get_all_file_path(r'data/双创wa项目接口文档.xlsx'),
                                      '二行情诗-发送弹幕', 1, 5, [2, 4, 7, 8, 9, 11, 12, 13, 14, 15, 16])
